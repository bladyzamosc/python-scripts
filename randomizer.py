import math
import random

import openpyxl

CONST_60 = 60

ONE_HUNDRED = 100

TOP_3_PERCENT = 3

TOP_3_MIN = 2

REST_2_PERCENT = 2

REST_2_MIN = 1

MAX_VALUES = 1000

CELL_L = 'L'


class Randomizer:
    def __init__(self, file):
        self.file = file
        self.wb = None
        self.ws = None

    def randomize(self):
        self.open_and_init()
        vals = self.extract_values()
        if len(vals) < TOP_3_MIN + REST_2_MIN:
            top_elements = vals.items()
            rest_elements = []
        else:
            rest_number = 1
            top_number = 2
            if len(vals) > CONST_60:
                top3 = math.ceil(len(vals) * TOP_3_PERCENT / ONE_HUNDRED)
                top_number = TOP_3_MIN if TOP_3_MIN > top3 else top3
                rest2 = math.ceil(len(vals) * REST_2_PERCENT / ONE_HUNDRED)
                rest_number = REST_2_MIN if REST_2_MIN > rest2 else rest2
            print(len(vals), " - t:", top_number, " - r:", rest_number)
            sorted_values = sorted(vals.items(), key=lambda x: x[1], reverse=True)
            top_elements = sorted_values[0:top_number]
            rest_elements = sorted_values[top_number:len(sorted_values)]
            print("top: ", top_elements)
            print("rest: ", rest_elements)
            random.shuffle(rest_elements)
            rest_elements = rest_elements[0:rest_number]

        self.copy_results(rest_elements, top_elements)

    def copy_results(self, rest_elements, top_elements):
        draw = self.wb.create_sheet("Losowanie")
        mc = self.ws.max_column + 1
        row_new = 1
        for t in top_elements:
            for j in range(1, mc):
                cell_from = self.ws.cell(row=t[0], column=j)
                draw.cell(row=row_new, column=j).value = cell_from.value
            draw.cell(row=row_new, column=(mc + 1)).value = "TOP"
            row_new = row_new + 1
        for t in rest_elements:
            for j in range(1, mc):
                cell_from = self.ws.cell(row=t[0], column=j)
                draw.cell(row=row_new, column=j).value = cell_from.value
            draw.cell(row=row_new, column=(mc + 1)).value = "RANDOM"
            row_new = row_new + 1
        self.wb.save(self.file)

    def open_and_init(self):
        self.wb = self.open_ws()
        self.ws = self.wb.active

    def open_ws(self):
        wb = openpyxl.load_workbook(self.file, data_only=True)
        return wb

    def extract_values(self):
        result = {}
        for cell in self.ws[CELL_L]:
            value = cell.value
            prev_val = self.ws.cell(row=cell.row, column=11).value
            if (isinstance(value, int) or isinstance(value, float)) and value is not None and prev_val != 'SUMA':
                result[cell.row] = value
            if cell.row > MAX_VALUES or prev_val == 'SUMA':
                print("Koniec")
                break
        return result
